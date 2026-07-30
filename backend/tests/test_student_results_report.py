"""The student-results report (SPECIFICATION.md §11) — data shape, PDF, Excel, and the route.

Mirrors ``tests/test_examination_office_report.py``'s conventions and fixture shapes, but is the
simpler sibling: no course/module grouping, no names anywhere in the output. §11 frames the
absence of names as a privacy property ("matches common practice of posting anonymized grade
lists"), so both the PDF and Excel renderer tests include an explicit negative assertion that no
name appears in the rendered output, not just a check that the right columns are present.

Only synthetic names are used throughout — never anything resembling a real registration PDF.
"""

from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
import pdfplumber
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.auth.passwords import hash_password
from app.grading.schema import GRADES
from app.models import (
    Exam,
    Exercise,
    ExercisePoints,
    GradeThreshold,
    Lecture,
    StudentRegistration,
    User,
)
from app.reports.student_results import (
    TEMPLATE_PATH,
    build_student_results_data,
    render_student_results_excel,
    render_student_results_pdf,
    student_results_filename,
)
from tests.conftest import INSTRUCTOR_PASSWORD, LoginHelper

#: Same schema/percentages as tests/test_examination_office_report.py's: on a 60-point exam every
#: percentage floors cleanly to itself (§7.2), giving point thresholds 57.0, 54.0, ..., 30.0 (best
#: to worst grade). 25 points is below every threshold ("nicht bestanden"); 60 points clears 1.0.
PERCENTAGES = ["95", "90", "85", "80", "75", "70", "65", "60", "55", "50"]

COURSE_A = "B.Sc. WiIng ET/IT"
MODULE_A = "Grundlagen der Informationstechnik (B.Sc. WiIng ET/IT)"
COURSE_B = "M.Sc. WiIng"
MODULE_B = "Grundlagen der Informationstechnik (M.Sc. WiIng)"


# --------------------------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------------------------


def _make_exam(session: Session, owner: User) -> Exam:
    lecture = Lecture(name="Grundlagen der Informationstechnik", owner_id=owner.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=owner.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    exam.exercises.append(Exercise(name="Aufgabe 1", max_points=Decimal(60), position=1))
    for grade, percentage in zip(GRADES, PERCENTAGES, strict=True):
        exam.grade_thresholds.append(GradeThreshold(grade=grade, percentage=Decimal(percentage)))
    session.add(exam)
    session.commit()
    return exam


def _registration(
    exam: Exam,
    *,
    matrikelnummer: str,
    nachname: str,
    vorname: str,
    course_code: str,
    module_title: str,
    attended: bool | None,
    points: Decimal | None = None,
    excluded: bool = False,
) -> StudentRegistration:
    registration = StudentRegistration(
        matrikelnummer=matrikelnummer,
        nachname=nachname,
        vorname=vorname,
        course_code=course_code,
        module_title=module_title,
        versuch=1,
        attended=attended,
        excluded=excluded,
    )
    if points is not None:
        exercise = exam.exercises[0]
        registration.exercise_points.append(ExercisePoints(exercise_id=exercise.id, points=points))
    exam.registrations.append(registration)
    return registration


@pytest.fixture
def complete_exam(session: Session, instructor_user: User) -> Exam:
    """A fully complete, fully configured exam covering all three §7.4 "Note" outcomes.

    - Musterfrau (course A): 60 points -> numeric grade "1,0".
    - Beispiel (course A): attended, 25 points -> "nicht bestanden".
    - Fehlt (course A): attended=False -> "n.e.".
    - Ausgeschlossen (course A): excluded -> omitted entirely.
    - Weit (course B): 60 points -> confirms course has no bearing on grouping/output at all.
    """
    exam = _make_exam(session, instructor_user)
    _registration(
        exam,
        matrikelnummer="1000010",
        nachname="Musterfrau",
        vorname="Erika",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
    )
    _registration(
        exam,
        matrikelnummer="1000005",
        nachname="Beispiel",
        vorname="Bruno",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(25),
    )
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Fehlt",
        vorname="Nils",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=False,
    )
    _registration(
        exam,
        matrikelnummer="1000099",
        nachname="Ausgeschlossen",
        vorname="Petra",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
        excluded=True,
    )
    _registration(
        exam,
        matrikelnummer="2000001",
        nachname="Weit",
        vorname="Wanda",
        course_code=COURSE_B,
        module_title=MODULE_B,
        attended=True,
        points=Decimal(60),
    )
    session.commit()
    return exam


@pytest.fixture
def incomplete_exam(session: Session, instructor_user: User) -> Exam:
    """A configured exam with one registration still missing its points (§8.1 blocks this)."""
    exam = _make_exam(session, instructor_user)
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Unvollstaendig",
        vorname="Uwe",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        # No points entered.
    )
    session.commit()
    return exam


@pytest.fixture
def unconfigured_exam(session: Session, instructor_user: User) -> Exam:
    """Otherwise-complete exam, but no grading schema configured at all (§8.1 blocks this too)."""
    lecture = Lecture(name="Grundlagen der Informationstechnik", owner_id=instructor_user.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=instructor_user.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    exam.exercises.append(Exercise(name="Aufgabe 1", max_points=Decimal(60), position=1))
    session.add(exam)
    session.commit()
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Musterfrau",
        vorname="Erika",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
    )
    session.commit()
    return exam


@pytest.fixture
def instructor_client(login: LoginHelper, instructor_user: User) -> TestClient:
    client, _token = login("dozentin", INSTRUCTOR_PASSWORD)
    return client


@pytest.fixture
def other_client(login: LoginHelper, session: Session) -> TestClient:
    session.add(
        User(
            username="dozent-b",
            password_hash=hash_password(INSTRUCTOR_PASSWORD),
            is_admin=False,
            is_active=True,
        )
    )
    session.commit()
    client, _token = login("dozent-b", INSTRUCTOR_PASSWORD)
    return client


def _pdf_text(pdf_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def _pdf_url(exam_id: int) -> str:
    return f"/api/exams/{exam_id}/reports/student-results/pdf"


def _excel_url(exam_id: int) -> str:
    return f"/api/exams/{exam_id}/reports/student-results/excel"


ALL_NAMES = ["Musterfrau", "Erika", "Beispiel", "Bruno", "Fehlt", "Nils", "Weit", "Wanda"]


# --------------------------------------------------------------------------------------------
# build_student_results_data — flat, sorted, no grouping, exclusion, Note values
# --------------------------------------------------------------------------------------------


def test_data_has_no_course_or_module_fields_at_all(complete_exam: Exam) -> None:
    """The one behavior most likely to accidentally regress: copying §10's grouping structure."""
    data = build_student_results_data(complete_exam)

    assert set(data.keys()) == {"lecture_name", "semester", "termin", "exam_date", "rows"}
    assert isinstance(data["rows"], list)
    for row in data["rows"]:
        assert set(row.keys()) == {"matrikelnummer", "note"}


def test_rows_are_flat_and_sorted_by_matrikelnummer_only(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)

    # Course B's student (matrikelnummer "2000001") is interleaved by matrikelnummer, not
    # grouped/sorted after course A as a section would be — proof there is no course grouping.
    assert [row["matrikelnummer"] for row in data["rows"]] == [
        "1000001",
        "1000005",
        "1000010",
        "2000001",
    ]


def test_excluded_registration_is_omitted_entirely(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)

    matrikelnummern = {row["matrikelnummer"] for row in data["rows"]}
    assert "1000099" not in matrikelnummern


def test_note_covers_all_three_outcomes(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)
    by_matrikelnummer = {row["matrikelnummer"]: row["note"] for row in data["rows"]}

    assert by_matrikelnummer["1000010"] == "1,0"  # Musterfrau
    assert by_matrikelnummer["1000005"] == "nicht bestanden"  # Beispiel
    assert by_matrikelnummer["1000001"] == "n.e."  # Fehlt


def test_lecture_and_exam_metadata(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)

    assert data["lecture_name"] == "Grundlagen der Informationstechnik"
    assert data["semester"] == "WiSe 23/24"
    assert data["termin"] == "1. Termin"
    assert data["exam_date"] is None


def test_zero_rows_when_no_registrations(session: Session, instructor_user: User) -> None:
    exam = _make_exam(session, instructor_user)

    data = build_student_results_data(exam)

    assert data["rows"] == []


def test_raises_runtime_error_if_schema_incomplete(session: Session, instructor_user: User) -> None:
    """This must never be reachable via HTTP — a caller bug, so a loud failure is correct."""
    lecture = Lecture(name="Ohne Schema", owner_id=instructor_user.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=instructor_user.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    session.add(exam)
    session.commit()

    with pytest.raises(RuntimeError):
        build_student_results_data(exam)


# --------------------------------------------------------------------------------------------
# render_student_results_pdf
# --------------------------------------------------------------------------------------------


def test_template_imports_no_package_at_all() -> None:
    code = "\n".join(
        line
        for line in TEMPLATE_PATH.read_text(encoding="utf-8").splitlines()
        if not line.lstrip().startswith("//")
    )
    assert "@preview" not in code
    assert "#import" not in code


def test_pdf_has_matrikelnummer_and_note_but_no_names(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)
    text = _pdf_text(render_student_results_pdf(data))

    for heading in ("Matr.-Nr.", "Note"):
        assert heading in text, heading

    for matrikelnummer in ("1000010", "1000005", "1000001", "2000001"):
        assert matrikelnummer in text, matrikelnummer
    assert "1000099" not in text  # excluded

    assert "1,0" in text
    assert "nicht bestanden" in text
    assert "n.e." in text

    # §11's core privacy property: no name — imported or excluded — appears anywhere.
    for name in ALL_NAMES:
        assert name not in text, name


def test_pdf_zero_rows_renders_a_valid_pdf_with_a_message(
    session: Session, instructor_user: User
) -> None:
    exam = _make_exam(session, instructor_user)
    data = build_student_results_data(exam)

    pdf_bytes = render_student_results_pdf(data)

    assert pdf_bytes.startswith(b"%PDF")
    text = _pdf_text(pdf_bytes)
    assert "keine Studierenden angemeldet" in text


# --------------------------------------------------------------------------------------------
# render_student_results_excel
# --------------------------------------------------------------------------------------------


def test_excel_has_two_columns_string_cells_and_no_names(complete_exam: Exam) -> None:
    data = build_student_results_data(complete_exam)
    workbook_bytes = render_student_results_excel(data)

    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    assert header == ["Matr.-Nr.", "Note"]
    assert sheet[1][0].font.bold is True
    assert sheet[1][1].font.bold is True

    rows = list(sheet.iter_rows(min_row=2, values_only=False))
    assert len(rows) == len(data["rows"])
    for row in rows:
        assert len(row) == 2

    by_matrikelnummer = {row[0].value: row for row in rows}
    assert isinstance(by_matrikelnummer["1000010"][0].value, str)  # Matr.-Nr. must never be numeric
    assert by_matrikelnummer["1000010"][1].value == "1,0"
    assert isinstance(by_matrikelnummer["1000010"][1].value, str)  # Note is mixed-type text

    assert by_matrikelnummer["1000005"][1].value == "nicht bestanden"
    assert isinstance(by_matrikelnummer["1000005"][1].value, str)

    assert by_matrikelnummer["1000001"][1].value == "n.e."
    assert isinstance(by_matrikelnummer["1000001"][1].value, str)

    assert "1000099" not in by_matrikelnummer  # excluded

    # §11's core privacy property: no name string appears anywhere in the sheet, not even as a
    # substring of some other cell — a plain membership check over cell values would miss e.g. a
    # cell containing "Musterfrau, Erika", so join everything into one haystack first.
    haystack = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row)
    for name in ALL_NAMES:
        assert name not in haystack, name


# --------------------------------------------------------------------------------------------
# Filename helper
# --------------------------------------------------------------------------------------------


def test_filename_sanitises_the_semester_slash_pdf(complete_exam: Exam) -> None:
    assert (
        student_results_filename(complete_exam, extension="pdf")
        == "Notenliste_WiSe_23-24_1._Termin.pdf"
    )


def test_filename_sanitises_the_semester_slash_excel(complete_exam: Exam) -> None:
    assert (
        student_results_filename(complete_exam, extension="xlsx")
        == "Notenliste_WiSe_23-24_1._Termin.xlsx"
    )


# --------------------------------------------------------------------------------------------
# The routes — §8.1 gate, ownership, media types, filenames
# --------------------------------------------------------------------------------------------


def test_pdf_route_returns_200_with_pdf_media_type_and_filename(
    instructor_client: TestClient, complete_exam: Exam
) -> None:
    response = instructor_client.get(_pdf_url(complete_exam.id))

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")
    disposition = response.headers["content-disposition"]
    assert 'filename="Notenliste_WiSe_23-24_1._Termin.pdf"' in disposition
    assert not disposition.rstrip('"').endswith(".xlsx.pdf")
    assert response.headers["cache-control"] == "no-store"


def test_excel_route_returns_200_with_excel_media_type_and_filename(
    instructor_client: TestClient, complete_exam: Exam
) -> None:
    response = instructor_client.get(_excel_url(complete_exam.id))

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert 'filename="Notenliste_WiSe_23-24_1._Termin.xlsx"' in disposition
    assert not disposition.rstrip('"').endswith(".xlsx.pdf")
    assert response.headers["cache-control"] == "no-store"

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.active is not None


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_incomplete_exam_returns_409_with_german_error(
    instructor_client: TestClient, incomplete_exam: Exam, url_fn
) -> None:
    response = instructor_client.get(url_fn(incomplete_exam.id))

    assert response.status_code == 409
    errors = response.json()["detail"]["errors"]
    assert any("unvollständig" in message for message in errors)


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_unconfigured_schema_returns_409_even_if_otherwise_complete(
    instructor_client: TestClient, unconfigured_exam: Exam, url_fn
) -> None:
    response = instructor_client.get(url_fn(unconfigured_exam.id))

    assert response.status_code == 409
    errors = response.json()["detail"]["errors"]
    assert any("Notenschema" in message for message in errors)


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_another_instructors_exam_is_404(
    other_client: TestClient, complete_exam: Exam, url_fn
) -> None:
    response = other_client.get(url_fn(complete_exam.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "Prüfung nicht gefunden."


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_unauthenticated_gets_401(client: TestClient, complete_exam: Exam, url_fn) -> None:
    response = client.get(url_fn(complete_exam.id))

    assert response.status_code == 401
