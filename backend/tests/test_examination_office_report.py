"""The examination-office report (SPECIFICATION.md §10) — data shape, PDF, Excel, and the route.

Follows ``tests/test_attendance_list.py``'s conventions: the interesting logic
(``build_examination_office_data``'s grouping/sort) is asserted on the plain dict, not scraped
back out of a rendered document, and the PDF/Excel renderers are exercised separately for the
"the byte output actually contains what it should" property that only parsing the format back
proves. HTTP-layer tests (the §8.1 gate returning ``409``, ownership returning ``404``) live in
this file too rather than a separate API test file, mirroring how ``test_attendance_list.py``
keeps its route tests alongside the data/render tests for the same report.

Only synthetic names are used throughout — never anything resembling a real registration PDF.
"""

from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
import pdfplumber
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.auth.passwords import hash_password
from app.grading.schema import GRADES
from app.models import (
    Exam,
    Exercise,
    ExercisePoints,
    GradeThreshold,
    Lecture,
    StudentRegistration,
    User,
)
from app.reports.examination_office import (
    TEMPLATE_PATH,
    build_examination_office_data,
    examination_office_filename,
    render_examination_office_excel,
    render_examination_office_pdf,
)
from tests.conftest import INSTRUCTOR_PASSWORD, LoginHelper

#: Same schema/percentages as ``tests/test_points_api.py``'s ``VALID_SCHEMA``: on a 60-point exam
#: every percentage floors cleanly to itself (§7.2), giving point thresholds 57.0, 54.0, 51.0,
#: 48.0, 45.0, 42.0, 39.0, 36.0, 33.0, 30.0 (best to worst grade). 25 points is below every
#: threshold ("nicht bestanden"); 60 points clears the best grade (1.0).
PERCENTAGES = ["95", "90", "85", "80", "75", "70", "65", "60", "55", "50"]

COURSE_A = "B.Sc. WiIng ET/IT"
MODULE_A = "Grundlagen der Informationstechnik (B.Sc. WiIng ET/IT)"
# Same course_code as COURSE_A/MODULE_A but a different module_title — the Kombinationsprüfung
# case: two course PDFs sharing a short code but carrying a different official module name.
MODULE_A_KOMBI = "Grundlagen der Informationstechnik, alte BPO (B.Sc. WiIng ET/IT)"
COURSE_B = "M.Sc. WiIng"
MODULE_B = "Grundlagen der Informationstechnik (M.Sc. WiIng)"


# --------------------------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------------------------


def _make_exam(session: Session, owner: User) -> Exam:
    lecture = Lecture(name="Grundlagen der Informationstechnik", owner_id=owner.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=owner.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    exam.exercises.append(Exercise(name="Aufgabe 1", max_points=Decimal(60), position=1))
    for grade, percentage in zip(GRADES, PERCENTAGES, strict=True):
        exam.grade_thresholds.append(GradeThreshold(grade=grade, percentage=Decimal(percentage)))
    session.add(exam)
    session.commit()
    return exam


def _registration(
    exam: Exam,
    *,
    matrikelnummer: str,
    nachname: str,
    vorname: str,
    course_code: str,
    module_title: str,
    attended: bool | None,
    points: Decimal | None = None,
    excluded: bool = False,
) -> StudentRegistration:
    registration = StudentRegistration(
        matrikelnummer=matrikelnummer,
        nachname=nachname,
        vorname=vorname,
        course_code=course_code,
        module_title=module_title,
        versuch=1,
        attended=attended,
        excluded=excluded,
    )
    if points is not None:
        exercise = exam.exercises[0]
        registration.exercise_points.append(ExercisePoints(exercise_id=exercise.id, points=points))
    exam.registrations.append(registration)
    return registration


@pytest.fixture
def complete_exam(session: Session, instructor_user: User) -> Exam:
    """A fully complete, fully configured exam covering all three §7.4 "Note" outcomes.

    - Musterfrau (course A / module A): 60 points -> numeric grade "1,0".
    - Beispiel (course A / module A): attended, 25 points -> "nicht bestanden".
    - Kombi (course A / module A **Kombi**, same course_code, different module_title): 60
      points -> its own section, not merged into module A.
    - Fehlt (course A / module A): attended=False -> "n.e.".
    - Ausgeschlossen (course A / module A): excluded -> omitted entirely.
    - Weit (course B / module B): 60 points -> exercises the multi-section case.
    """
    exam = _make_exam(session, instructor_user)
    _registration(
        exam,
        matrikelnummer="1000010",
        nachname="Musterfrau",
        vorname="Erika",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
    )
    _registration(
        exam,
        matrikelnummer="1000005",
        nachname="Beispiel",
        vorname="Bruno",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(25),
    )
    _registration(
        exam,
        matrikelnummer="1000020",
        nachname="Kombi",
        vorname="Klara",
        course_code=COURSE_A,
        module_title=MODULE_A_KOMBI,
        attended=True,
        points=Decimal(60),
    )
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Fehlt",
        vorname="Nicht",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=False,
    )
    _registration(
        exam,
        matrikelnummer="1000099",
        nachname="Ausgeschlossen",
        vorname="Petra",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
        excluded=True,
    )
    _registration(
        exam,
        matrikelnummer="2000001",
        nachname="Weit",
        vorname="Wanda",
        course_code=COURSE_B,
        module_title=MODULE_B,
        attended=True,
        points=Decimal(60),
    )
    session.commit()
    return exam


@pytest.fixture
def incomplete_exam(session: Session, instructor_user: User) -> Exam:
    """A configured exam with one registration still missing its points (§8.1 blocks this)."""
    exam = _make_exam(session, instructor_user)
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Unvollstaendig",
        vorname="Uwe",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        # No points entered.
    )
    session.commit()
    return exam


@pytest.fixture
def unconfigured_exam(session: Session, instructor_user: User) -> Exam:
    """Otherwise-complete exam, but no grading schema configured at all (§8.1 blocks this too)."""
    lecture = Lecture(name="Grundlagen der Informationstechnik", owner_id=instructor_user.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=instructor_user.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    exam.exercises.append(Exercise(name="Aufgabe 1", max_points=Decimal(60), position=1))
    session.add(exam)
    session.commit()
    _registration(
        exam,
        matrikelnummer="1000001",
        nachname="Musterfrau",
        vorname="Erika",
        course_code=COURSE_A,
        module_title=MODULE_A,
        attended=True,
        points=Decimal(60),
    )
    session.commit()
    return exam


@pytest.fixture
def instructor_client(login: LoginHelper, instructor_user: User) -> TestClient:
    client, _token = login("dozentin", INSTRUCTOR_PASSWORD)
    return client


@pytest.fixture
def other_client(login: LoginHelper, session: Session) -> TestClient:
    session.add(
        User(
            username="dozent-b",
            password_hash=hash_password(INSTRUCTOR_PASSWORD),
            is_admin=False,
            is_active=True,
        )
    )
    session.commit()
    client, _token = login("dozent-b", INSTRUCTOR_PASSWORD)
    return client


def _pdf_text(pdf_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def _pdf_url(exam_id: int) -> str:
    return f"/api/exams/{exam_id}/reports/examination-office/pdf"


def _excel_url(exam_id: int) -> str:
    return f"/api/exams/{exam_id}/reports/examination-office/excel"


# --------------------------------------------------------------------------------------------
# build_examination_office_data — grouping, sorting, exclusion, Note values
# --------------------------------------------------------------------------------------------


def test_same_course_code_different_module_title_are_two_separate_sections(
    complete_exam: Exam,
) -> None:
    """The one behavior most likely to get silently wrong: grouping must be on the *pair*."""
    data = build_examination_office_data(complete_exam)

    course_a_sections = [s for s in data["sections"] if s["course_code"] == COURSE_A]
    assert len(course_a_sections) == 2

    module_titles = {s["module_title"] for s in course_a_sections}
    assert module_titles == {MODULE_A, MODULE_A_KOMBI}

    kombi_section = next(s for s in course_a_sections if s["module_title"] == MODULE_A_KOMBI)
    assert [r["nachname"] for r in kombi_section["rows"]] == ["Kombi"]

    module_a_section = next(s for s in course_a_sections if s["module_title"] == MODULE_A)
    assert {r["nachname"] for r in module_a_section["rows"]} == {"Musterfrau", "Beispiel", "Fehlt"}


def test_sections_and_rows_are_sorted(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)

    # Sections: course B ("M.Sc. WiIng") sorts after course A ("B.Sc. WiIng ET/IT"); within
    # course A, module_title MODULE_A ("Grundlagen...") sorts before MODULE_A_KOMBI
    # ("Grundlagen..., alte BPO...").
    section_keys = [(s["course_code"], s["module_title"]) for s in data["sections"]]
    assert section_keys == [
        (COURSE_A, MODULE_A),
        (COURSE_A, MODULE_A_KOMBI),
        (COURSE_B, MODULE_B),
    ]

    module_a_section = data["sections"][0]
    # Sorted by Matrikelnummer (plain string sort): "1000001" < "1000005" < "1000010".
    assert [r["matrikelnummer"] for r in module_a_section["rows"]] == [
        "1000001",
        "1000005",
        "1000010",
    ]


def test_excluded_registration_is_omitted_entirely(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)

    all_names = {r["nachname"] for s in data["sections"] for r in s["rows"]}
    assert "Ausgeschlossen" not in all_names


def test_note_covers_all_three_outcomes(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)
    by_name = {r["nachname"]: r["note"] for s in data["sections"] for r in s["rows"]}

    assert by_name["Musterfrau"] == "1,0"
    assert by_name["Beispiel"] == "nicht bestanden"
    assert by_name["Fehlt"] == "n.e."


def test_lecture_and_exam_metadata(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)

    assert data["lecture_name"] == "Grundlagen der Informationstechnik"
    assert data["semester"] == "WiSe 23/24"
    assert data["termin"] == "1. Termin"
    assert data["exam_date"] is None


def test_zero_sections_when_no_registrations(session: Session, instructor_user: User) -> None:
    exam = _make_exam(session, instructor_user)

    data = build_examination_office_data(exam)

    assert data["sections"] == []


def test_raises_runtime_error_if_schema_incomplete(session: Session, instructor_user: User) -> None:
    """This must never be reachable via HTTP — a caller bug, so a loud failure is correct."""
    lecture = Lecture(name="Ohne Schema", owner_id=instructor_user.id)
    session.add(lecture)
    session.flush()
    exam = Exam(
        lecture_id=lecture.id,
        owner_id=instructor_user.id,
        semester="WiSe 23/24",
        termin="1. Termin",
    )
    session.add(exam)
    session.commit()

    with pytest.raises(RuntimeError):
        build_examination_office_data(exam)


# --------------------------------------------------------------------------------------------
# render_examination_office_pdf
# --------------------------------------------------------------------------------------------


def test_template_imports_no_package_at_all() -> None:
    code = "\n".join(
        line
        for line in TEMPLATE_PATH.read_text(encoding="utf-8").splitlines()
        if not line.lstrip().startswith("//")
    )
    assert "@preview" not in code
    assert "#import" not in code


def test_pdf_shows_module_title_as_section_heading_and_note_values(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)
    text = _pdf_text(render_examination_office_pdf(data))

    assert MODULE_A in text
    assert MODULE_A_KOMBI in text
    assert MODULE_B in text
    for heading in ("Matr.-Nr.", "Nachname", "Vorname", "Note"):
        assert heading in text, heading

    assert "nicht bestanden" in text
    assert "n.e." in text
    assert "Ausgeschlossen" not in text


def test_pdf_zero_sections_renders_a_valid_pdf_with_a_message(
    session: Session, instructor_user: User
) -> None:
    exam = _make_exam(session, instructor_user)
    data = build_examination_office_data(exam)

    pdf_bytes = render_examination_office_pdf(data)

    assert pdf_bytes.startswith(b"%PDF")
    text = _pdf_text(pdf_bytes)
    assert "keine Studierenden angemeldet" in text


# --------------------------------------------------------------------------------------------
# render_examination_office_excel
# --------------------------------------------------------------------------------------------


def test_excel_has_flat_rows_with_module_title_column_and_string_cells(complete_exam: Exam) -> None:
    data = build_examination_office_data(complete_exam)
    workbook_bytes = render_examination_office_excel(data)

    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    assert header == ["Matr.-Nr.", "Nachname", "Vorname", "Note", "Modultitel"]
    assert sheet[1][0].font.bold is True

    rows = list(sheet.iter_rows(min_row=2, values_only=False))
    # Same row count/order as the PDF (flattened sections, in order) so a spot-check lines up.
    total_rows = sum(len(s["rows"]) for s in data["sections"])
    assert len(rows) == total_rows

    by_name = {row[1].value: row for row in rows}
    musterfrau = by_name["Musterfrau"]
    assert musterfrau[0].value == "1000010"
    assert isinstance(musterfrau[0].value, str)  # Matr.-Nr. must never be numeric
    assert musterfrau[3].value == "1,0"
    assert isinstance(musterfrau[3].value, str)  # Note is mixed-type text
    assert musterfrau[4].value == MODULE_A

    beispiel = by_name["Beispiel"]
    assert beispiel[3].value == "nicht bestanden"
    assert isinstance(beispiel[3].value, str)

    fehlt = by_name["Fehlt"]
    assert fehlt[3].value == "n.e."
    assert isinstance(fehlt[3].value, str)

    kombi = by_name["Kombi"]
    assert kombi[4].value == MODULE_A_KOMBI

    assert "Ausgeschlossen" not in by_name


# --------------------------------------------------------------------------------------------
# Filename helper
# --------------------------------------------------------------------------------------------


def test_filename_sanitises_the_semester_slash_pdf(complete_exam: Exam) -> None:
    assert (
        examination_office_filename(complete_exam, extension="pdf")
        == "Pruefungsamt_WiSe_23-24_1._Termin.pdf"
    )


def test_filename_sanitises_the_semester_slash_excel(complete_exam: Exam) -> None:
    assert (
        examination_office_filename(complete_exam, extension="xlsx")
        == "Pruefungsamt_WiSe_23-24_1._Termin.xlsx"
    )


# --------------------------------------------------------------------------------------------
# The routes — §8.1 gate, ownership, media types, filenames
# --------------------------------------------------------------------------------------------


def test_pdf_route_returns_200_with_pdf_media_type_and_filename(
    instructor_client: TestClient, complete_exam: Exam
) -> None:
    response = instructor_client.get(_pdf_url(complete_exam.id))

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")
    disposition = response.headers["content-disposition"]
    assert 'filename="Pruefungsamt_WiSe_23-24_1._Termin.pdf"' in disposition
    assert not disposition.rstrip('"').endswith(".xlsx.pdf")
    assert response.headers["cache-control"] == "no-store"


def test_excel_route_returns_200_with_excel_media_type_and_filename(
    instructor_client: TestClient, complete_exam: Exam
) -> None:
    response = instructor_client.get(_excel_url(complete_exam.id))

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert 'filename="Pruefungsamt_WiSe_23-24_1._Termin.xlsx"' in disposition
    # The bug that was just fixed: an ASCII fallback ending in "...xlsx.pdf".
    assert not disposition.rstrip('"').endswith(".xlsx.pdf")
    assert response.headers["cache-control"] == "no-store"

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.active is not None


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_incomplete_exam_returns_409_with_german_error(
    instructor_client: TestClient, incomplete_exam: Exam, url_fn
) -> None:
    response = instructor_client.get(url_fn(incomplete_exam.id))

    assert response.status_code == 409
    errors = response.json()["detail"]["errors"]
    assert any("unvollständig" in message for message in errors)


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_unconfigured_schema_returns_409_even_if_otherwise_complete(
    instructor_client: TestClient, unconfigured_exam: Exam, url_fn
) -> None:
    response = instructor_client.get(url_fn(unconfigured_exam.id))

    assert response.status_code == 409
    errors = response.json()["detail"]["errors"]
    assert any("Notenschema" in message for message in errors)


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_another_instructors_exam_is_404(
    other_client: TestClient, complete_exam: Exam, url_fn
) -> None:
    response = other_client.get(url_fn(complete_exam.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "Prüfung nicht gefunden."


@pytest.mark.parametrize("url_fn", [_pdf_url, _excel_url])
def test_unauthenticated_gets_401(client: TestClient, complete_exam: Exam, url_fn) -> None:
    response = client.get(url_fn(complete_exam.id))

    assert response.status_code == 401
